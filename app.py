from gc import collect
import time
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import logging
import os
from tkinter import messagebox

path = "./FT_a_procesar"
contenido = os.listdir(path)
fichas = []

for ficha in contenido:
    if os.path.isfile(os.path.join(path, ficha)) and ficha.endswith(".xlsm"):
        fichas.append(ficha)
print(fichas)

logging.basicConfig(
    filename="app.txt",
    level=logging.INFO,
    format="%(asctime)s:%(levelname)s:%(message)s",
)

# read/write Excel

""" Download chromedriver """
# chrome://version/
# https://chromedriver.storage.googleapis.com/index.html
# pip install openpyxl


logging.info("...Iniciando...")

#'openpyxl.worksheet.merge.MergedCellRange', MergedCellRange


# driver_service = Service(executable_path="./selenium-driver/chromedriver.exe")
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()
# URL
driver.get("http://grisinobftestfinal:8001/maer/")  # cambiar url aca


class Login:
    def __init__(self, user, Falseword):
        self.user = user
        self.Falseword = Falseword

    def login(self):
        try:
            logging.info("Iniciando sesion..")
            login_user = WebDriverWait(driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.ID, "ext-comp-1002")
                )
            )

            login_user.send_keys(self.user)

            Falseword_user = WebDriverWait(driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.ID, "ext-comp-1004")
                )
            )
            Falseword_user.send_keys(self.Falseword)

            login_btn = driver.find_element(By.ID, "ext-gen31")
            login_btn.click()
            logging.info("Entrando...")
        except (Exception) as error_excepction:
            logging.warning("Error: ", error_excepction)


class LoadFile:
    def __init__(self, fichas):
        self.fichas = fichas
        logging.info("Cargando file...")

    def loop(self, rango, lista):
        # Iterar por cod de color
        for cod in rango:
            for i in cod:
                if i.value is not None:
                    lista.append(i.value)

    def split_cod_color(self, cod_color):
        if cod_color is not None:
            return cod_color.split("-", 1)[1]
        else:
            return ""

    def loop_cod_color(self, rango_cod_color, lista_cod_color, celda, rango_str, ws):
        # comprobar si la celda esta mergeada para elegir un color o varios
        merged_cell_ranges = ws.merged_cells.ranges
        merged_cell_ranges_list = list(map(str, merged_cell_ranges))
        merged_cells = []

        for merged_range in merged_cell_ranges_list:
            merged_cells.append(merged_range)

        if rango_str in merged_cells:
            print("Cell is Merged!")
            merged_cell_cod_color = self.split_cod_color(celda.value)
            lista_cod_color.append(merged_cell_cod_color)
            print("109: ", lista_cod_color)
            return True
        else:
            print("Cell is not merged")
            for cod in rango_cod_color:
                for i in cod:
                    if i.value is not None:
                        single_cod_color = self.split_cod_color(i.value)
                        lista_cod_color.append(single_cod_color)
                        print("116: ", lista_cod_color)
            return False

    def comprobar_y_cargar(
        self,
        actions,
        descripcion_validacion,
        talles,
        lista_cod_color,
        cantidad_insumo_confeccion,
        insumo_confeccion,
        lista_colores,
        isCombined,
        agregar_insumo,
    ):
        if "TALLE" in descripcion_validacion:
            logging.info("Cargnado insumo por talle...")
            # Por cada talle cargar...}
            for talle in talles:
                time.sleep(1)
                logging.info("talles disponibles: ", talle)
                for idx, i in enumerate(lista_cod_color):
                    self.load_insumo_por_talle(
                        actions,
                        insumo_confeccion,
                        i,
                        cantidad_insumo_confeccion,
                        talle,
                        lista_colores[idx],
                        isCombined,
                        agregar_insumo,
                    )
        else:
            for idx, i in enumerate(lista_cod_color):
                self.load_insumo2(
                    actions,
                    insumo_confeccion,
                    i,
                    cantidad_insumo_confeccion,
                    lista_colores[idx],
                    isCombined,
                    agregar_insumo,
                )
            logging.info("Carga de insumo por talle finalizada...")

    def load_insumo(self, actions, insumo, cod_color_insumo, cantidad):
        if insumo is not None:

            logging.info(f"Cargando el insumo {insumo}")
            time.sleep(2)
            actions.send_keys(insumo + "." + cod_color_insumo)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(2)
            actions.send_keys(cantidad)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
        else:
            actions.send_keys(Keys.ESCAPE)
            actions.perform()
            actions.send_keys(Keys.ESCAPE)
            actions.perform()
            logging.info(f"Carga de insumo {insumo} finalizada")

    def load_insumo2(
        self, actions, insumo, i, cantidad, color, isCombined, agregar_insumo
    ):
        time.sleep(1)
        logging.info(f"Cargando inusmo por talle: {insumo}")
        agregar_insumo.click()
        time.sleep(2)
        actions.send_keys(Keys.TAB)
        actions.perform()
        time.sleep(1)
        print(f"Cargando cod de color: {i}")
        actions.send_keys(insumo + "." + i)
        actions.perform()
        time.sleep(2)
        actions.send_keys(Keys.ENTER)
        actions.perform()
        time.sleep(2)
        actions.send_keys(Keys.TAB)
        actions.perform()
        time.sleep(2)
        actions.send_keys(cantidad)
        actions.perform()
        time.sleep(2)
        actions.send_keys(Keys.ENTER)
        actions.perform()
        time.sleep(2)
        if isCombined:
            actions.send_keys(Keys.ENTER)
            print(f"Cargando color Todos")
            actions.perform()
        else:
            print("Color a cargar: ", color)
            actions.send_keys(color)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ARROW_DOWN)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
        time.sleep(2)
        actions.send_keys("Todos")
        actions.perform()
        time.sleep(1)
        actions.send_keys(Keys.ESCAPE)
        actions.perform()

    def load_insumo_por_talle(
        self,
        actions,
        insumo,
        cod_color_insumo,
        cantidad,
        talle,
        color,
        isCombined,
        agregar_insumo,
    ):
        if insumo is not None:
            logging.info(f"Cargando inusmo por talle: {insumo}")
            time.sleep(2)
            agregar_insumo.click()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(1)
            actions.send_keys(insumo + "." + cod_color_insumo + "." + talle)
            actions.perform()
            time.sleep(3)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(2)
            actions.send_keys(cantidad)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            time.sleep(1)
            actions.perform()
            time.sleep(3)
            if isCombined:
                actions.send_keys(Keys.ENTER)
                print(f"Cargando color 'Todos' para insumo {insumo}")
                actions.perform()
            else:
                print("Color a cargar: ", color + "para insumo: ", insumo)
                actions.send_keys(color)
                time.sleep(1)
                actions.perform()
                time.sleep(1)
                actions.send_keys(Keys.ENTER)
            time.sleep(2)
            if (
                talle == "6"
                or talle == "8"
                or talle == "10"
                or talle == "12"
                or talle == "14"
            ):
                actions.send_keys(talle + " (GRISINO C3-C4 NUMEROS)")
                time.sleep(1)
                actions.perform()
            elif talle == "1" or talle == "2" or talle == "3" or talle == "4":
                actions.send_keys(talle + " (GRISINO C1-C2 NUMEROS)")
                time.sleep(1)
                actions.perform()
                time.sleep(1)
            else:
                actions.send_keys(talle + " (GRISINO C0 ABCDE)")
                time.sleep(1)
                actions.perform()
            time.sleep(2)

    def load_new(self):
        try:
            time.sleep(2)
            btn_produccion = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.XPATH,
                        "/html/body/div[1]/div[2]/div/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[2]/em/button",
                    )
                )
            )
            btn_produccion.click()
            time.sleep(2)
            btn_ficha_tecnica = driver.find_element(
                By.XPATH,
                "//span[contains(text(),'Fichas Técnicas')]",
            )
            btn_ficha_tecnica.click()

            btn_ficha_tecnica2 = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (By.ID, "menuPrincipalProducciónFichas TécnicasFichas Técnicas")
                )
            )
            btn_ficha_tecnica2.click()

            btn_maxim_ft = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.CLASS_NAME,
                        "x-tool-maximize",
                    )
                )
            )
            btn_maxim_ft.click()
            btn_add_new = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.XPATH,
                        "//button[contains(text(),'Agregar')]",
                    )
                )
            )
            time.sleep(2)
            btn_add_new.click()
            logging.info("reading excel..")

            for index, ficha in enumerate(self.fichas):

                logging.info(f"Cargando ficha: {ficha}")
                wb = load_workbook(f"./FT_a_procesar/{ficha}", data_only=True)
                ws = wb.active

                time.sleep(40)

                input_coleccion = driver.find_element(
                    By.XPATH,
                    "//input[@id='ext-comp-1254']",
                )
                time.sleep(5)
                input_coleccion.click()
                actions = ActionChains(driver)
                coleccion_parte1 = ws["B1"].value
                coleccion_parte2 = ws["B2"].value
                coleccion_parte3 = ws["L5"].value
                time.sleep(2)
                coleccion = (
                    f"{coleccion_parte1} {coleccion_parte2[0]} - {coleccion_parte3}"
                )
                print("Coleccion: ", coleccion)
                time.sleep(2)
                actions.send_keys(coleccion)
                time.sleep(2)
                actions.send_keys(Keys.ENTER)
                time.sleep(2)
                actions.perform()
                time.sleep(3)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(3)
                producto = ws["B2"].value
                time.sleep(3)
                actions.send_keys(producto)
                time.sleep(2)
                actions.perform()
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(3)
                actions.send_keys(Keys.TAB)
                actions.perform()
                actions.send_keys(Keys.TAB)
                actions.perform()
                molde = ws["T2"].value
                actions.send_keys(molde)
                actions.perform()
                time.sleep(4)

                btn_add_rule = driver.find_element(
                    By.XPATH,
                    "//table[@id='ext-comp-1281']/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(5)
                logging.info("Agregando regla - telas corte")
                btn_add_rule.click()
                time.sleep(2)
                actions.send_keys("100 - CORTE")
                time.sleep(2)
                actions.perform()
                time.sleep(1)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(1)
                actions.send_keys(Keys.ESCAPE)
                actions.perform()
                time.sleep(3)

                nueva_entrada = driver.find_element(
                    By.XPATH,
                    "//div[@id='ext-comp-1276']/div/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(5)
                nueva_entrada.click()
                time.sleep(2)

                if index == 0:
                    agregar_insumo = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1324']/tbody/tr[2]/td[2]/em/button",
                    )
                elif index == 1:
                    agregar_insumo = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1705']/tbody/tr[2]/td[2]/em/button",
                    )

                time.sleep(2)
                agregar_insumo.click()
                logging.info("Agregando insumos telas")
                time.sleep(1)
                actions.send_keys(Keys.TAB)
                actions.perform()

                insumo_1 = ws["I6"].value

                cod_color_inusmo = ws["L7"].value

                cod_color_insumo2 = ws["L9"].value

                # Cantidad
                cantidad_insumo_1 = str(ws["J6"].value)
                cantidad_insumo_2 = ws["J8"].value

                time.sleep(2)
                self.load_insumo(
                    actions,
                    insumo_1,
                    self.split_cod_color(cod_color_inusmo),
                    cantidad_insumo_1,
                )
                time.sleep(2)
                insumo_2 = ws["I8"].value
                insumo_3 = ws["I10"].value

                cod_color_insumo3 = ws["L11"].value

                cantidad_insumo_3 = ws["J10"].value
                insumo_4 = ws["I12"].value
                insumo_5 = ws["I14"].value
                insumo_6 = ws["I16"].value
                cod_color_insumo4 = ws["L13"].value

                cod_color_insumo5 = ws["L15"].value
                cod_color_insumo6 = ws["L17"].value

                cantidad_insumo_4 = str(ws["J12"].value)
                cantidad_insumo_5 = str(ws["J14"].value)
                cantidad_insumo_6 = str(ws["J16"].value)

                # Si insumo existe.. agregar otro
                # Se puede hacer una fx decoradora -----------------------------------------------------------
                if insumo_2 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_2,
                        self.split_cod_color(cod_color_insumo2),
                        cantidad_insumo_2,
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_3 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_3,
                        self.split_cod_color(cod_color_insumo3),
                        cantidad_insumo_3,
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_4 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_4,
                        self.split_cod_color(cod_color_insumo4),
                        cantidad_insumo_4,
                    )

                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_5 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_5,
                        self.split_cod_color(cod_color_insumo5),
                        cantidad_insumo_5,
                    )

                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_6 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_6,
                        self.split_cod_color(cod_color_insumo6),
                        cantidad_insumo_6,
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                # ----------------------------------------------     300 - bordados      -----------------------------------------------------------------------------------------------
                logging.info("Comenzando carga de bordados...")
                time.sleep(1)
                btn_add_rule.click()
                time.sleep(1)
                proceso_corte = WebDriverWait(driver, 35).until(
                    expected_conditions.presence_of_element_located(
                        (
                            By.XPATH,
                            "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[2]/div/input",
                        )
                    )
                )
                descripcion_validacion_1_bordado = ws["B22"].value
                time.sleep(2)
                if descripcion_validacion_1_bordado == "BORDADO":
                    proceso_corte.send_keys("300 - BORDADO")
                else:
                    proceso_corte.send_keys("400 - ESTAMPADO")

                time.sleep(2)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(2)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                time.sleep(2)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(2)
                actions.send_keys(Keys.ESCAPE)
                actions.perform()

                time.sleep(3)

                # /html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button
                nueva_entrada_bordado = WebDriverWait(driver, 35).until(
                    expected_conditions.presence_of_element_located(
                        (
                            By.XPATH,
                            "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                        )
                    )
                )
                time.sleep(3)
                nueva_entrada_bordado.click()
                time.sleep(1)

                insumo_bordado_1 = ws["I22"].value
                insumo_bordado_2 = ws["I24"].value
                insumo_confeccion_1 = ws["I26"].value
                insumo_confeccion_2 = ws["I28"].value
                insumo_confeccion_3 = ws["I30"].value
                insumo_confeccion_4 = ws["I32"].value
                insumo_confeccion_5 = ws["I34"].value
                insumo_elastico_1 = ws["I36"].value
                insumo_hilo_1 = ws["I74"].value
                insumo_conf_1 = ws["I76"].value
                insumo_lav_1 = ws["I77"].value
                insumo_avios_de_lav_conf_1 = ws["I79"].value
                insumo_ilustrador_1 = ws["I104"].value

                isCombined1_bordado = ws["L22"]
                isCombined2_bordado = ws["L24"]
                isCombined1 = ws["L26"]
                isCombined2 = ws["L28"]
                isCombined3 = ws["L30"]
                isCombined4 = ws["L32"]
                isCombined5 = ws["L34"]

                cantidad_insumo_bordado_1 = ws["J22"].value
                cantidad_insumo_bordado_2 = ws["J24"].value
                cantidad_insumo_confeccion_1 = ws["J26"].value
                cantidad_insumo_confeccion_2 = ws["J28"].value
                cantidad_insumo_confeccion_3 = ws["J30"].value
                cantidad_insumo_confeccion_4 = ws["J32"].value
                cantidad_insumo_confeccion_5 = ws["J34"].value

                # CANTIDADES SEGUNDO COD DE PROD
                cantidad_insumo_bordado_1_cod_2 = ws["K22"].value
                cantidad_insumo_bordado_2_cod_2 = ws["K24"].value
                cantidad_insumo_confeccion_1_cod_2 = ws["K26"].value
                cantidad_insumo_confeccion_2_cod_2 = ws["K28"].value
                cantidad_insumo_confeccion_3_cod_2 = ws["K30"].value
                cantidad_insumo_confeccion_4_cod_2 = ws["K32"].value
                cantidad_insumo_confeccion_5_cod_2 = ws["K34"].value

                descripcion_validacion_1_bordado = ws["B22"].value
                descripcion_validacion_2_bordado = ws["B24"].value

                descripcion_validacion_1 = ws["B26"].value
                descripcion_validacion_2 = ws["B28"].value
                descripcion_validacion_3 = ws["B30"].value
                descripcion_validacion_4 = ws["B32"].value
                descripcion_validacion_5 = ws["B34"].value

                rango_cod_color_str_bordado1 = "L22:U22"
                rango_cod_color_str_bordado2 = "L24:U24"
                rango_cod_color1_str = "L26:U26"
                rango_cod_color2_str = "L28:U28"
                rango_cod_color3_str = "L30:U30"
                rango_cod_color4_str = "L32:U32"
                rango_cod_color5_str = "L34:U34"

                rango_cod_color_bordado1 = ws["L22":"U22"]
                rango_cod_color_bordado2 = ws["L24":"U24"]
                rango_cod_color1 = ws["L26":"U26"]
                rango_cod_color2 = ws["L28":"U28"]
                rango_cod_color3 = ws["L30":"U30"]
                rango_cod_color4 = ws["L32":"U32"]
                rango_cod_color5 = ws["L34":"U34"]

                lista_cod_color_1_bordado = []
                lista_cod_color_2_bordado = []
                lista_cod_color_1 = []
                lista_cod_color_2 = []
                lista_cod_color_3 = []
                lista_cod_color_4 = []
                lista_cod_color_5 = []
                lista_cod_color_6 = []

                time.sleep(3)

                """ Elasticos vars """

                lista_cod_color_elastico_1 = []
                rango_cod_color_elastico_1 = ws["L36":"U36"]
                isCombined_elastico_1 = ws["L36"]
                rango_cod_color1_str_elastico = "L36:U36"
                descripcion_validacion_1_elastico = ws["B36"].value
                cantidad_elastico_1 = ws["J36"].value
                cantidad_elastico_1_cod_2 = ws["K36"].value
                cod_color_elastico_1 = ws["L36"].value

                """ Hilos """
                lista_cod_color_hilo_1 = []
                rango_cod_color_hilo_1 = ws["L74":"U74"]
                isCombined_hilo_1 = ws["L74"]
                rango_cod_color1_str_hilo = "L74:U74"
                descripcion_validacion_1_hilo = ws["B74"].value
                cantidad_hilo_1 = ws["J74"].value
                cantidad_hilo_1_cod_2 = ws["K74"].value
                cod_color_hilo_1 = ws["L74"].value

                """ Confeccion """
                lista_cod_color_confeccion_1 = []
                rango_cod_color_confeccion_1 = ws["L64":"U26"]
                isCombined_confeccion_1 = ws["L76"]
                rango_cod_color1_str_confeccion = "L64:U76"
                descripcion_validacion_1_confeccion = ws["B76"].value
                cantidad_confeccion_1 = ws["J76"].value
                cantidad_confeccion_1_cod_2 = ws["K76"].value
                cod_color_confeccion_1 = ws["L76"].value

                """ Lavado """
                cod_color_lav_1 = ws["L77"].value
                cantidad_lav_1 = ws["J77"].value

                """ Avios de lav y confecc """
                cod_color_avios_lav_con_1 = ws["L79"].value
                cantidad_avios_lav_con_1 = ws["J79"].value
                cantidad_avios_lav_con_1_cod_2 = ws["K79"].value

                """ Plancha """
                insumo_plancha_1 = ws["K79"].value

                """ GTS """
                cantidad_gts_1 = ws["J04"].value
                cantidad_gts_1_cod_2 = ws["K104"].value

                time.sleep(3)
            
                lista_colores = []
                rango_colores = ws["L4":"T4"]

                talles_value = ws["P2"].value
                talles = [str(x) for x in talles_value.split(" - ")]
                talles2 = []

                talles_value_2 = ws["P3"].value
                """ Talles segundo cod de prod """
                if talles_value_2 is not None:
                    talles2 = [str(x) for x in talles_value_2.split(" - ")]
                else:
                    False

                # LOOPS
                self.loop(rango_colores, lista_colores)

                # MERGED CELLS
                combinado_1_bordado = self.loop_cod_color(
                    rango_cod_color_bordado1,
                    lista_cod_color_1_bordado,
                    isCombined1_bordado,
                    rango_cod_color_str_bordado1,
                    ws,
                )

                combinado_2_bordado = self.loop_cod_color(
                    rango_cod_color_bordado2,
                    lista_cod_color_2_bordado,
                    isCombined2_bordado,
                    rango_cod_color_str_bordado2,
                    ws,
                )
                time.sleep(3)
                if insumo_bordado_1 is not None:
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_1_bordado,
                        talles,
                        lista_cod_color_1_bordado,
                        cantidad_insumo_bordado_1,
                        insumo_bordado_1,
                        lista_colores,
                        combinado_1_bordado,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_bordado_1} terminada")
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                time.sleep(2)

                if insumo_bordado_2 is not None:
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_2_bordado,
                        talles,
                        lista_cod_color_2_bordado,
                        cantidad_insumo_bordado_2,
                        insumo_bordado_2,
                        lista_colores,
                        combinado_2_bordado,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_bordado_2} terminada")
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                # ------------------------------------------------------200  confeccion ----------------------------------------------------------------------------
                logging.info("Comenzando carga de confeccion...")
                time.sleep(1)
                btn_add_rule.click()
                time.sleep(1)
                actions.send_keys("600 - PREPARACION P/ TALLER")
                time.sleep(1)
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                actions.send_keys(Keys.ENTER)
                actions.perform()
                actions.send_keys(Keys.ESCAPE)
                actions.perform()

                time.sleep(3)
                nueva_entrada_prep_taller = WebDriverWait(driver, 35).until(
                    expected_conditions.presence_of_element_located(
                        (
                            By.XPATH,
                            "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[3]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                        )
                    )
                )
                nueva_entrada_prep_taller.click()
                time.sleep(1)

                combinado_1 = self.loop_cod_color(
                    rango_cod_color1,
                    lista_cod_color_1,
                    isCombined1,
                    rango_cod_color1_str,
                    ws,
                )

                combinado_2 = self.loop_cod_color(
                    rango_cod_color2,
                    lista_cod_color_2,
                    isCombined2,
                    rango_cod_color2_str,
                    ws,
                )
                combinado_3 = self.loop_cod_color(
                    rango_cod_color3,
                    lista_cod_color_3,
                    isCombined3,
                    rango_cod_color3_str,
                    ws,
                )
                combinado_4 = self.loop_cod_color(
                    rango_cod_color4,
                    lista_cod_color_4,
                    isCombined4,
                    rango_cod_color4_str,
                    ws,
                )
                combinado_5 = self.loop_cod_color(
                    rango_cod_color5,
                    lista_cod_color_5,
                    isCombined5,
                    rango_cod_color5_str,
                    ws,
                )
                # FX
                time.sleep(2)
                if insumo_confeccion_1 is not None:
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_1,
                        talles,
                        lista_cod_color_1,
                        cantidad_insumo_confeccion_1,
                        insumo_confeccion_1,
                        lista_colores,
                        combinado_1,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_confeccion_1} terminada")

                if insumo_confeccion_2 is not None:
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_2,
                        talles,
                        lista_cod_color_2,
                        cantidad_insumo_confeccion_2,
                        insumo_confeccion_2,
                        lista_colores,
                        combinado_2,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_confeccion_2} terminada")

                time.sleep(2)

                if insumo_confeccion_3 is not None:
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_3,
                        talles,
                        lista_cod_color_3,
                        cantidad_insumo_confeccion_3,
                        insumo_confeccion_3,
                        lista_colores,
                        combinado_3,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_confeccion_3} terminada")

                if insumo_confeccion_4 is not None:
                    print(f"Paso por el {insumo_confeccion_4}")
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_4,
                        talles,
                        lista_cod_color_4,
                        cantidad_insumo_confeccion_4,
                        insumo_confeccion_4,
                        lista_colores,
                        combinado_4,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_confeccion_4} terminada")

                if insumo_confeccion_5 is not None:
                    print(f"Paso por el {insumo_confeccion_5}")
                    self.comprobar_y_cargar(
                        actions,
                        descripcion_validacion_5,
                        talles,
                        lista_cod_color_5,
                        cantidad_insumo_confeccion_5,
                        insumo_confeccion_5,
                        lista_colores,
                        combinado_5,
                        agregar_insumo,
                    )
                    logging.info(f"Carga de inusmo: {insumo_confeccion_5} termiada")

                """ Elasticos """
                time.sleep(3)
                if insumo_elastico_1 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_elastico_1,
                        self.split_cod_color(cod_color_elastico_1),
                        cantidad_elastico_1,
                    )
                    logging.info(f"Carga de inusmo: {insumo_elastico_1} termiada")

                """ Hilos """

                print(f"insumo hilo : {insumo_hilo_1}")
                if insumo_hilo_1 is not None:
                    time.sleep(2)
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_hilo_1,
                        self.split_cod_color(cod_color_hilo_1),
                        cantidad_hilo_1,
                    )
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info(f"Carga de inusmo: {insumo_hilo_1} termiada")

                """ Confeccion """
                if insumo_conf_1 is not None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_conf_1,
                        self.split_cod_color(cod_color_confeccion_1),
                        cantidad_confeccion_1,
                    )
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info(f"Carga de inusmo: {insumo_conf_1} termiada")

                actions.send_keys(Keys.ESCAPE)
                actions.perform()
                actions.send_keys(Keys.ESCAPE)
                actions.perform()

                """ Lavado """
                if insumo_lav_1 is not None:
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("520 - LAVADO")
                    time.sleep(1)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)
                    nueva_entrada_lavado = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[4]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                            )
                        )
                    )
                    nueva_entrada_lavado.click()
                    time.sleep(1)
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_lav_1,
                        self.split_cod_color(cod_color_lav_1),
                        cantidad_lav_1,
                    )
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info(f"Carga de inusmo: {insumo_lav_1} termiada")

                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                """ Avios de lav y confeccion """
                if insumo_avios_de_lav_conf_1 is not None:
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("480 - APROBACION LAVADO")
                    time.sleep(1)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)
                    nueva_entrada_avios_lav_con = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[4]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                            )
                        )
                    )
                    nueva_entrada_avios_lav_con.click()
                    time.sleep(1)
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_avios_de_lav_conf_1,
                        self.split_cod_color(cod_color_avios_lav_con_1),
                        cantidad_avios_lav_con_1,
                    )
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info(
                        f"Carga de inusmo: {insumo_avios_de_lav_conf_1} termiada"
                    )

                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                """ Plancha """
                lista_insumos_plancha = []
                cantidades_plancha = []
                rango_insumo_plancha = ws["I94":"I103"]
                rango_cantidades_insumo_plancha = ws["J94":"J103"]
                self.loop(rango_insumo_plancha, lista_insumos_plancha)
                self.loop(rango_cantidades_insumo_plancha, cantidades_plancha)
                print(f"Insumos en la lista de plancha: {lista_insumos_plancha}")

                if lista_insumos_plancha[0] is not None:
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("750 - PLANCHA")
                    time.sleep(1)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)

                    nueva_entrada_plancha = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[5]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                            )
                        )
                    )   
                    time.sleep(3)
                    nueva_entrada_plancha.click()
                    time.sleep(1)
                    for i in lista_insumos_plancha:
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        for c in cantidades_plancha:
                            self.load_insumo(actions, i, "SC.U", c)
                            time.sleep(1)
                            actions.send_keys(Keys.ESCAPE)
                            actions.perform()
                            break
                        time.sleep(2)

                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                """ Ilustrador """
                if insumo_ilustrador_1 is not None:
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("800 - PRODUCTO TERMINADO")
                    time.sleep(1)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)
                    nueva_entrada_gts = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[6]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                            )
                        )
                    )
                    nueva_entrada_gts.click()
                    time.sleep(1)
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(1)
                    self.load_insumo(
                        actions,
                        insumo_ilustrador_1,
                        "SC.U", cantidad_gts_1
                    )
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info(
                        f"Carga de inusmo: {insumo_ilustrador_1} termiada"
                    )

                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()                
                # --------------------------------------------------- GUARDA PRIMERA PARTE -----------------------------------------------------
                time.sleep(2)
                btn_guardar = driver.find_element(
                    By.XPATH,
                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[2]/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(2)
                btn_guardar.click()
                time.sleep(5)
                btn_si = driver.find_element(
                    By.XPATH, "//button[contains(text(),'Sí')]"
                )
                time.sleep(5)
                btn_si.click()

                btn_ok = driver.find_element(
                    By.XPATH, "//button[contains(text(),'OK')]"
                )
                time.sleep(5)
                btn_ok.click()
                logging.info(f"Ficha:{ficha} con cod de producto: {coleccion} guardada")
                time.sleep(5)
                btn_close = driver.find_element(
                    By.XPATH,
                    "//div[@id='ext-comp-1505']/div/div/div/div/div",
                )
                time.sleep(2)
                btn_close.click()
                time.sleep(2)
                actions.send_keys(Keys.ESCAPE)
                actions.perform()

                """ ------------------------------------------------------------  SEGUNDO COD DE PRODUCTO  ---------------------------------------------------- """

                cod_art_2 = ws["B3"].value

                if cod_art_2 is not None:
                    logging.info(
                        f"Cargando ficha:{ficha} con segundo codigo de producto: {cod_art_2}"
                    )
                    input_coleccion2 = driver.find_element(
                        By.XPATH,
                        "//*[@id='ext-comp-1254']",
                    )
                    actions = ActionChains(driver)
                    coleccion = ws["G1"].value
                    time.sleep(3)
                    input_coleccion2.send_keys(coleccion)
                    time.sleep(3)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(cod_art_2)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    molde = ws["T2"].value
                    actions.send_keys(molde)
                    actions.perform()

                    time.sleep(2)
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("100 - CORTE")
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(2)
                    logging.info("Agregando entrada")
                    nueva_entrada2 = driver.find_element(
                        By.XPATH,
                        "//div[@id='ext-comp-1276']/div/div[2]/div/div/div[2]/div/div/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                    )
                    nueva_entrada2.click()
                    time.sleep(2)

                    agregar_insumo2 = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1527']/tbody/tr[2]/td[2]/em/button",
                    )
                    agregar_insumo2.click()
                    logging.info("Cargando insumos...")
                    time.sleep(4)
                    actions.send_keys(Keys.TAB)
                    actions.perform()

                    insumo_1 = ws["I6"].value
                    cantidad_insumo_1 = str(ws["K6"].value)
                    cantidad_insumo_2 = str(ws["K8"].value)

                    time.sleep(2)
                    self.load_insumo(
                        actions,
                        insumo_1,
                        self.split_cod_color(cod_color_inusmo),
                        cantidad_insumo_1,
                    )
                    time.sleep(2)
                    insumo_2 = ws["I8"].value
                    insumo_3 = ws["I10"].value
                    color_insumo3 = ws["L11"].value
                    cantidad_insumo_3 = ws["K10"].value
                    insumo_4 = ws["I12"].value
                    insumo_5 = ws["I14"].value
                    insumo_6 = ws["I16"].value
                    color_insumo4 = ws["N5"].value
                    # XTA004
                    color_insumo5 = ws["N5"].value
                    # XTD001
                    color_insumo6 = ws["N5"].value
                    cantidad_insumo_4 = str(ws["K12"].value)
                    cantidad_insumo_5 = str(ws["K14"].value)
                    cantidad_insumo_6 = str(ws["K16"].value)

                    # Si insumo existe.. agregar otro
                    # Se puede hacer una fx decoradora -----------------------------------------------------------
                    if insumo_2 is not None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_2, cod_color_insumo2, cantidad_insumo_2
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_3 is not None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_3, color_insumo3, cantidad_insumo_3
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info(f"Carga de insumo {insumo_3} finalizada")

                    time.sleep(2)

                    if insumo_4 is not None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_4, color_insumo4, cantidad_insumo_4
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_5 is not None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_5, color_insumo5, cantidad_insumo_5
                        )

                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_6 is not None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_6, color_insumo6, cantidad_insumo_6
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(4)

                    # ----------------------------------------------     300 - bordados      -----------------------------------------------------------------------------------------------
                    logging.info("Comenzando carga de bordados...")
                    time.sleep(1)
                    btn_add_rule.click()
                    time.sleep(1)
                    proceso_corte = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[2]/div/input",
                            )
                        )
                    )

                    descripcion_validacion_1_bordado = ws["B22"].value
                    time.sleep(2)
                    if descripcion_validacion_1_bordado == "BORDADO":
                        proceso_corte.send_keys("300 - BORDADO")
                    else:
                        proceso_corte.send_keys("400 - ESTAMPADO")

                    time.sleep(2)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)

                    # /html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button
                    nueva_entrada_bordado.click()
                    time.sleep(1)

                    if insumo_bordado_1 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_1_bordado,
                            talles2,
                            lista_cod_color_1_bordado,
                            cantidad_insumo_bordado_1_cod_2,
                            insumo_bordado_1,
                            lista_colores,
                            combinado_1_bordado,
                            agregar_insumo,
                        )
                        logging.info(f"Carga de inusmo: {insumo_bordado_1} terminada")
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()

                    time.sleep(2)

                    if insumo_bordado_2 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_2_bordado,
                            talles2,
                            lista_cod_color_2_bordado,
                            cantidad_insumo_bordado_2_cod_2,
                            insumo_bordado_2,
                            lista_colores,
                            combinado_2_bordado,
                            agregar_insumo,
                        )
                        logging.info(f"Carga de inusmo: {insumo_bordado_2} terminada")
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                    # ------------------------------------------------------  600 - prep para taller ----------------------------------------------------------------------------
                    logging.info("Comenzando carga de confeccion...")
                    time.sleep(1)
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("600 - PREPARACION P/ TALLER")
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(2)
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(3)

                    nueva_entrada_prep_taller = WebDriverWait(driver, 35).until(
                        expected_conditions.presence_of_element_located(
                            (
                                By.XPATH,
                                "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[3]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                            )
                        )
                    )
                    nueva_entrada_prep_taller.click()
                    time.sleep(1)

                    # /html/body/div[4]/div[2]/div/div/div/div[1]/div[9]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button
                    nueva_entrada_bordado2 = driver.find_element(
                        By.XPATH,
                        "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[2]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(5)
                    nueva_entrada_bordado2.click()
                    time.sleep(1)

                    # Variables de insumos aca

                    lista_colores = []
                    rango_colores = ws["L4":"T4"]

                    # LOOPS
                    self.loop(rango_colores, lista_colores)

                    # FX
                    time.sleep(1)

                    if insumo_confeccion_1 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_1,
                            talles2,
                            lista_cod_color_1,
                            cantidad_insumo_confeccion_1_cod_2,
                            insumo_confeccion_1,
                            lista_colores,
                            combinado_1,
                            agregar_insumo2,
                        )
                        logging.info(
                            f"Carga de inusmo: {insumo_confeccion_1} terminada"
                        )

                    time.sleep(2)

                    if insumo_confeccion_2 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_2,
                            talles2,
                            lista_cod_color_2,
                            cantidad_insumo_confeccion_2_cod_2,
                            insumo_confeccion_2,
                            lista_colores,
                            combinado_2,
                            agregar_insumo2,
                        )
                        logging.info(
                            f"Carga de inusmo: {insumo_confeccion_2} terminada"
                        )

                    time.sleep(2)

                    if insumo_confeccion_3 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_3,
                            talles2,
                            lista_cod_color_3,
                            cantidad_insumo_confeccion_3_cod_2,
                            insumo_confeccion_3,
                            lista_colores,
                            combinado_3,
                            agregar_insumo2,
                        )
                        logging.info(
                            f"Carga de inusmo: {insumo_confeccion_3} terminada"
                        )

                    if insumo_confeccion_4 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_4,
                            talles2,
                            lista_cod_color_4,
                            cantidad_insumo_confeccion_4_cod_2,
                            insumo_confeccion_4,
                            lista_colores,
                            combinado_4,
                            agregar_insumo,
                        )
                        logging.info(
                            f"Carga de inusmo: {insumo_confeccion_4} terminada"
                        )
                    if insumo_confeccion_5 is not None:
                        self.comprobar_y_cargar(
                            actions,
                            descripcion_validacion_5,
                            talles2,
                            lista_cod_color_5,
                            cantidad_insumo_confeccion_5_cod_2,
                            insumo_confeccion_5,
                            lista_colores,
                            combinado_5,
                            agregar_insumo,
                        )
                        logging.info(
                            f"Carga de inusmo: {insumo_confeccion_5} termi6ada"
                        )

                    """ Elasticos """
                    time.sleep(3)
                    if insumo_elastico_1 is not None:
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        self.load_insumo(
                            actions,
                            insumo_elastico_1,
                            self.split_cod_color(cod_color_elastico_1),
                            cantidad_elastico_1_cod_2,
                        )
                        logging.info(f"Carga de inusmo: {insumo_elastico_1} termiada")

                    """ Hilos """
                    if insumo_hilo_1 is not None:
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        self.load_insumo(
                            actions,
                            insumo_hilo_1,
                            self.split_cod_color(cod_color_hilo_1),
                            cantidad_hilo_1_cod_2,
                        )
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info(f"Carga de inusmo: {insumo_hilo_1} termiada")

                    """ Confeccion """
                    if insumo_conf_1 is not None:
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        self.load_insumo(
                            actions,
                            insumo_conf_1,
                            self.split_cod_color(cod_color_confeccion_1),
                            cantidad_confeccion_1_cod_2,
                        )
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info(f"Carga de inusmo: {insumo_conf_1} termiada")

                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    """ Lavado """
                    if insumo_lav_1 is not None:
                        btn_add_rule.click()
                        time.sleep(1)
                        actions.send_keys("520 - LAVADO")
                        time.sleep(1)
                        actions.send_keys(Keys.ARROW_DOWN)
                        actions.perform()
                        actions.send_keys(Keys.ARROW_DOWN)
                        actions.perform()
                        actions.send_keys(Keys.ENTER)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()

                        time.sleep(3)
                        nueva_entrada_lavado = WebDriverWait(driver, 35).until(
                            expected_conditions.presence_of_element_located(
                                (
                                    By.XPATH,
                                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[4]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                                )
                            )
                        )
                        nueva_entrada_lavado.click()
                        time.sleep(1)
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        self.load_insumo(
                            actions,
                            insumo_lav_1,
                            self.split_cod_color(cod_color_lav_1),
                            cantidad_lav_1,
                        )
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info(f"Carga de inusmo: {insumo_lav_1} termiada")

                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()

                    """ Avios de lavado y conf """
                    if insumo_avios_de_lav_conf_1 is not None:
                        btn_add_rule.click()
                        time.sleep(1)
                        actions.send_keys("480 - APROBACION LAVADO")
                        time.sleep(1)
                        actions.send_keys(Keys.ARROW_DOWN)
                        actions.perform()
                        actions.send_keys(Keys.ARROW_DOWN)
                        actions.perform()
                        actions.send_keys(Keys.ENTER)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()

                        time.sleep(3)
                        nueva_entrada_avios_lav_con = WebDriverWait(driver, 35).until(
                            expected_conditions.presence_of_element_located(
                                (
                                    By.XPATH,
                                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[4]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                                )
                            )
                        )
                        nueva_entrada_avios_lav_con.click()
                        time.sleep(1)
                        agregar_insumo.click()
                        actions.send_keys(Keys.TAB)
                        actions.perform()
                        time.sleep(1)
                        self.load_insumo(
                            actions,
                            insumo_avios_de_lav_conf_1,
                            self.split_cod_color(cod_color_avios_lav_con_1),
                            cantidad_avios_lav_con_1_cod_2,
                        )
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info(
                            f"Carga de inusmo: {insumo_avios_de_lav_conf_1} termiada"
                        )

                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()

                        """ Plancha """
                        lista_insumos_plancha = []
                        cantidades_plancha = []
                        rango_insumo_plancha = ws["I94":"I103"]
                        rango_cantidades_insumo_plancha = ws["K94":"K103"]
                        self.loop(rango_insumo_plancha, lista_insumos_plancha)
                        self.loop(rango_cantidades_insumo_plancha, cantidades_plancha)
                        print(f"Insumos en la lista de plancha: {lista_insumos_plancha}")

                        if lista_insumos_plancha[0] is not None:
                            btn_add_rule.click()
                            time.sleep(1)
                            actions.send_keys("750 - PLANCHA")
                            time.sleep(1)
                            actions.send_keys(Keys.ARROW_DOWN)
                            actions.perform()
                            actions.send_keys(Keys.ARROW_DOWN)
                            actions.perform()
                            actions.send_keys(Keys.ENTER)
                            actions.perform()
                            actions.send_keys(Keys.ESCAPE)
                            actions.perform()

                            time.sleep(3)

                            nueva_entrada_plancha = WebDriverWait(driver, 35).until(
                                expected_conditions.presence_of_element_located(
                                    (
                                        By.XPATH,
                                        "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div[5]/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                                    )
                                )
                            )   
                            time.sleep(3)
                            nueva_entrada_plancha.click()
                            time.sleep(1)
                            for i in lista_insumos_plancha:
                                agregar_insumo.click()
                                actions.send_keys(Keys.TAB)
                                actions.perform()
                                time.sleep(1)
                                for c in cantidades_plancha:
                                    self.load_insumo(actions, i, "SC.U", c)
                                    f"Carga de inusmo: {i} termiada"
                                    time.sleep(1)
                                    actions.send_keys(Keys.ESCAPE)
                                    actions.perform()
                                    break
                                time.sleep(2)

                            actions.send_keys(Keys.ESCAPE)
                            actions.perform()
                            actions.send_keys(Keys.ESCAPE)
                            actions.perform()

                    # --------------------------------------------------- GUARDA SEGUNDA PARTE -----------------------------------------------------
                    time.sleep(3)
                    btn_guardar2 = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1304']/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(2)
                    btn_guardar2.click()
                    time.sleep(2)
                    btn_si2 = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1480']/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(3)
                    btn_si2.click()
                    logging.info("Ficha Guardada")
                    time.sleep(2)
                    btn_ok2 = driver.find_element(
                        By.XPATH,
                        "//button[contains(text(),'OK')]",
                    )
                    time.sleep(2)
                    btn_ok2.click()
                    time.sleep(2)
                    btn_close2 = driver.find_element(
                        By.XPATH, "//div[@id='ext-comp-1547']/div/div/div/div/div"
                    )
                    btn_close2.click()
                    time.sleep(2)
                    logging.info(f"Ficha: ${ficha} cargada exitosamente")
                else:
                    False
            # -------------------------------------------------- --------------------------------------------------------------------------

        except (Exception) as error_excepction:
            logging.info(f"Error: {error_excepction} , {Exception}")
            print(error_excepction)


# Credenciales
log = Login("RobotPRD", "Robot123")
log.login()


# Cargar Nueva ficha
load = LoadFile(fichas)
load.load_new()
